from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from python_calamine import load_workbook as load_calamine_workbook
except ImportError:  # pragma: no cover - fallback for older machines
    load_calamine_workbook = None

try:
    from openpyxl import load_workbook as load_openpyxl_workbook
except ImportError:  # pragma: no cover - calamine-only read path
    load_openpyxl_workbook = None


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = PROJECT_ROOT / "data" / "採購提醒設定.xlsx"
DEFAULT_OUTPUT = (
    PROJECT_ROOT
    / "public"
    / "phoenixes-film-inventory"
    / "purchase-alert-data.js"
)

WAREHOUSE_COLUMNS = [
    ("台北倉", "台北觀察線", "台北採購線"),
    ("台中倉", "台中觀察線", "台中採購線"),
    ("台南倉", "臺南觀察線", "臺南採購線"),
    ("高雄倉", "高雄觀察線", "高雄採購線"),
    ("欣凱倉", "欣凱觀察線", "欣凱採購線"),
]


def parse_number(value: Any) -> float | int | None:
    if value in (None, ""):
        return None
    try:
        number = float(str(value).replace(",", "").strip())
    except ValueError:
        return None
    if number <= 0:
        return None
    return int(number) if number.is_integer() else number


def parse_optional_number(value: Any) -> float | int | None:
    if value in (None, ""):
        return None
    try:
        number = float(str(value).replace(",", "").strip())
    except ValueError:
        return None
    if number < 0:
        return None
    return int(number) if number.is_integer() else number


def enabled_value(value: Any) -> bool:
    return str(value or "Y").strip().upper() not in {"N", "NO", "FALSE", "0"}


def latest_source(path: Path) -> Path:
    if path.is_file():
        return path
    files = [
        item
        for item in path.glob("*.xlsx")
        if item.is_file() and not item.name.startswith("~$")
    ]
    if not files:
        raise FileNotFoundError(f"No purchase alert .xlsx file found in {path}")
    return max(files, key=lambda item: item.stat().st_mtime)


def row_dict(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    return {
        str(header or "").strip(): values[index] if index < len(values) else None
        for index, header in enumerate(headers)
    }


def read_rows(path: Path) -> tuple[list[str], list[list[Any]], str]:
    if load_calamine_workbook is not None:
        workbook = load_calamine_workbook(str(path))
        sheet = (
            workbook.get_sheet_by_name("採購提醒設定")
            if "採購提醒設定" in workbook.sheet_names
            else workbook.get_sheet_by_index(0)
        )
        rows = sheet.to_python()
        if not rows:
            return [], [], "python-calamine"
        return [str(value or "").strip() for value in rows[0]], rows[1:], "python-calamine"

    if load_openpyxl_workbook is None:
        raise RuntimeError("Install python-calamine or openpyxl to read Excel files.")

    workbook = load_openpyxl_workbook(path, data_only=True)
    sheet = workbook["採購提醒設定"] if "採購提醒設定" in workbook.sheetnames else workbook.worksheets[0]
    headers = [
        str(sheet.cell(1, column).value or "").strip()
        for column in range(1, sheet.max_column + 1)
    ]
    rows = [
        [sheet.cell(row, column).value for column in range(1, sheet.max_column + 1)]
        for row in range(2, sheet.max_row + 1)
    ]
    return headers, rows, "openpyxl"


def build_settings(path: Path) -> dict[str, Any]:
    headers, rows, engine = read_rows(path)
    settings: dict[str, dict[str, Any]] = {}
    configured_count = 0
    enabled_count = 0

    for values in rows:
        row = row_dict(headers, values)
        name = str(row.get("品名") or "").strip()
        if not name:
            continue

        enabled = enabled_value(row.get("啟用(Y/N)"))
        thresholds: dict[str, dict[str, float | int]] = {}

        for warehouse, watch_column, order_column in WAREHOUSE_COLUMNS:
            watch = parse_number(row.get(watch_column))
            order = parse_number(row.get(order_column))
            if watch is not None or order is not None:
                thresholds[warehouse] = {}
                if watch is not None:
                    thresholds[warehouse]["watch"] = watch
                if order is not None:
                    thresholds[warehouse]["order"] = order

        total_watch = parse_number(row.get("五倉合計觀察線"))
        total_order = parse_number(row.get("五倉合計採購線"))
        total = {}
        if total_watch is not None:
            total["watch"] = total_watch
        if total_order is not None:
            total["order"] = total_order

        if thresholds or total:
            configured_count += 1
            if enabled:
                enabled_count += 1

        list_price = parse_optional_number(row.get("牌價"))
        setting = {
            "enabled": enabled,
            "category": row.get("分類") or "",
            "series": row.get("系列") or "",
            "widthMm": row.get("寬度") or "",
            "thresholds": thresholds,
            "total": total,
            "note": row.get("備註") or "",
        }
        if list_price is not None:
            setting["listPrice"] = list_price
        settings[name] = setting

    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "source": {
            "path": str(path),
            "filename": path.name,
            "engine": engine,
            "lastModified": datetime.fromtimestamp(path.stat().st_mtime).isoformat(
                timespec="seconds"
            ),
            "sizeBytes": path.stat().st_size,
        },
        "settings": settings,
        "summary": {
            "itemCount": len(settings),
            "configuredCount": configured_count,
            "enabledConfiguredCount": enabled_count,
        },
    }


def write_js(data: dict[str, Any], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    output.write_text(
        "window.PURCHASE_ALERT_SETTINGS = " + payload + ";\n",
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Build purchase alert settings data.")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    source = latest_source(args.source)
    data = build_settings(source)
    write_js(data, args.output)

    print(f"Source: {source}")
    print(f"Output: {args.output}")
    print(f"Engine: {data['source']['engine']}")
    print(f"Configured: {data['summary']['configuredCount']}")
    print(f"Enabled configured: {data['summary']['enabledConfiguredCount']}")


if __name__ == "__main__":
    main()
