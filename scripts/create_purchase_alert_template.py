from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "採購提醒設定.xlsx"

sys.path.insert(0, str(PROJECT_ROOT / "scripts"))
from build_dashboard import CATEGORY_RULES, derive_width_mm  # noqa: E402


WAREHOUSES = ["台北倉", "台中倉", "臺南倉", "高雄倉", "欣凱倉"]

HEADERS = [
    "分類",
    "系列",
    "品名",
    "寬度",
    "啟用(Y/N)",
    "台北觀察線",
    "台北採購線",
    "台中觀察線",
    "台中採購線",
    "臺南觀察線",
    "臺南採購線",
    "高雄觀察線",
    "高雄採購線",
    "欣凱觀察線",
    "欣凱採購線",
    "五倉合計觀察線",
    "五倉合計採購線",
    "備註",
]

INSTRUCTION_ROWS = [
    ("用途", "這份 Excel 用來設定每個品項、每個倉庫的採購提醒門檻。"),
    ("啟用(Y/N)", "Y=啟用提醒；N=不檢查該品項。"),
    ("觀察線", "實際庫存低於此數量時，網頁可列入「觀察中」。空白或 0 表示不檢查。"),
    ("採購線", "實際庫存低於此數量時，網頁可列入「需採購」。空白或 0 表示不檢查。"),
    ("分倉邏輯", "各倉獨立判斷；例如台北很多，但臺南低於門檻，仍會警示臺南。"),
    ("合計欄位", "五倉合計觀察線/採購線為選填；需要看總量時才填。"),
    ("建議", "先填重點品項即可，沒有填門檻的欄位不會跳提醒。"),
]


def iter_items() -> list[dict[str, str | int | None]]:
    items: list[dict[str, str | int | None]] = []
    for category in CATEGORY_RULES:
        for series in category["series"]:
            for item_name in series["items"]:
                items.append(
                    {
                        "category": category["category"],
                        "series": series["name"],
                        "name": item_name,
                        "width": derive_width_mm(item_name),
                    }
                )
    return items


def set_column_widths(ws) -> None:
    widths = {
        "A": 18,
        "B": 16,
        "C": 38,
        "D": 10,
        "E": 12,
        "F": 13,
        "G": 13,
        "H": 13,
        "I": 13,
        "J": 13,
        "K": 13,
        "L": 13,
        "M": 13,
        "N": 13,
        "O": 13,
        "P": 17,
        "Q": 17,
        "R": 28,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def style_settings_sheet(ws, row_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="1D1D1F")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=row_count, min_col=1, max_col=len(HEADERS)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in range(2, row_count + 1):
        ws[f"E{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center")
        for column in range(6, 18):
            ws.cell(row=row, column=column).alignment = Alignment(horizontal="right", vertical="center")

    enabled_validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
    ws.add_data_validation(enabled_validation)
    enabled_validation.add(f"E2:E{row_count}")

    number_validation = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
    )
    ws.add_data_validation(number_validation)
    number_validation.add(f"F2:Q{row_count}")

    observation_fill = PatternFill("solid", fgColor="FFF4D6")
    purchase_fill = PatternFill("solid", fgColor="FFE2E2")
    disabled_fill = PatternFill("solid", fgColor="F0F0F2")

    ws.conditional_formatting.add(
        f"A2:R{row_count}",
        FormulaRule(formula=['$E2="N"'], fill=disabled_fill),
    )

    for column in ["F", "H", "J", "L", "N", "P"]:
        ws.conditional_formatting.add(
            f"{column}2:{column}{row_count}",
            FormulaRule(formula=[f"AND(${column}2<>\"\",${column}2>0)"], fill=observation_fill),
        )

    for column in ["G", "I", "K", "M", "O", "Q"]:
        ws.conditional_formatting.add(
            f"{column}2:{column}{row_count}",
            FormulaRule(formula=[f"AND(${column}2<>\"\",${column}2>0)"], fill=purchase_fill),
        )

    ws.freeze_panes = "F2"
    ws.auto_filter.ref = f"A1:R{row_count}"
    ws.sheet_view.showGridLines = False

    ws["F1"].comment = Comment("觀察線：低於此數量列入觀察中。", "Codex")
    ws["G1"].comment = Comment("採購線：低於此數量列入需採購。採購線通常小於或等於觀察線。", "Codex")
    ws["P1"].comment = Comment("合計欄位是選填；分倉警示仍會各自判斷。", "Codex")


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "採購提醒設定"
    ws.append(HEADERS)

    for item in iter_items():
        ws.append(
            [
                item["category"],
                item["series"],
                item["name"],
                item["width"] or "",
                "Y",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )

    set_column_widths(ws)
    style_settings_sheet(ws, ws.max_row)

    instructions = wb.create_sheet("填寫說明")
    instructions.append(["項目", "說明"])
    for row in INSTRUCTION_ROWS:
        instructions.append(row)
    instructions.append(["建立時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    instructions.column_dimensions["A"].width = 22
    instructions.column_dimensions["B"].width = 88
    instructions.sheet_view.showGridLines = False
    for cell in instructions[1]:
        cell.fill = PatternFill("solid", fgColor="1D1D1F")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in instructions.iter_rows(min_row=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    plan = wb.create_sheet("網頁規劃")
    plan_rows = [
        ["功能", "建議做法"],
        ["採購觀察篩選", "新增一鍵篩選：全部 / 觀察中 / 需採購。"],
        ["分倉判斷", "任一倉庫低於該品項該倉門檻，就讓品項進入提醒清單。"],
        ["提醒顯示", "品項旁顯示簡短標籤，例如「臺南需採購」「台中觀察中」。"],
        ["優先級", "需採購優先於觀察中；同一品項多倉低於門檻時合併顯示。"],
        ["合計門檻", "如果有填五倉合計門檻，再額外判斷總量；未填就忽略。"],
        ["不自動算採購量", "第一版只提示狀態，不建議採購數量，避免忽略廠商基本量與倉庫空間。"],
    ]
    for row in plan_rows:
        plan.append(row)
    plan.column_dimensions["A"].width = 22
    plan.column_dimensions["B"].width = 88
    plan.sheet_view.showGridLines = False
    for cell in plan[1]:
        cell.fill = PatternFill("solid", fgColor="1D1D1F")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in plan.iter_rows(min_row=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description="Create purchase alert settings template.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    output = args.output
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(output)
    print(output)


if __name__ == "__main__":
    main()
